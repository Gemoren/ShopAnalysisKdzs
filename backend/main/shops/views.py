import os
import threading
import uuid
from datetime import datetime

import openpyxl
from django.conf import settings
from django.http import JsonResponse
from django.shortcuts import render
from .models import SysShop
from user.models import ImportTask
from django.core.paginator import Paginator
from django.views import View


# Create your views here.
class UploadFile(View):
    def post(self, request):
        print("request:", request)
        file = request.FILES.get('file')
        print("file:", file)

        if file:
            file_name = file.name
            suffixName = file_name[file_name.rfind("."):]
            new_file_name = datetime.now().strftime('%Y%m%d%H%M%S') + suffixName
            file_path = str(settings.MEDIA_ROOT) + "\\shopsFile\\" + new_file_name

            if not os.path.exists(str(settings.MEDIA_ROOT) + "\\shopsFile\\"):
                os.makedirs(str(settings.MEDIA_ROOT) + "\\shopsFile\\")

            print("file_path:", file_path)

            try:
                # 保存文件
                with open(file_path, 'wb') as f:
                    for chunk in file.chunks():
                        f.write(chunk)

                # 创建导入任务
                task_id = str(uuid.uuid4())
                task = ImportTask.objects.create(
                    task_id=task_id,
                    file_name=file_name,
                    file_path=file_path,
                    status='pending'
                )

                # 启动后台线程处理文件
                thread = threading.Thread(target=process_excel_file, args=(task_id, file_path))
                thread.daemon = True
                thread.start()

                return JsonResponse({
                    'code': 200,
                    'message': '文件上传成功，正在后台处理',
                    'task_id': task_id,
                    'file_name': new_file_name
                })

            except Exception as e:
                return JsonResponse({'code': 500, 'errorInfo': f'文件上传失败: {str(e)}'})

        return JsonResponse({'code': 400, 'errorInfo': '请上传文件'})


# 异步处理Excel文件的函数
def process_excel_file(task_id, file_path):
    try:
        print(f"开始处理任务: {task_id}")

        # 更新任务状态为处理中
        task = ImportTask.objects.get(task_id=task_id)
        task.status = 'processing'
        task.save()

        # 打开表格
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        # 获取表头（第一行）
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        print("表头:", headers)

        # 创建verbose_name到字段名的映射
        verbose_name_map = {
            '规格ID': 'spec_id',
            '规格名称': 'spec_name',
            '规格别名': 'spec_alias',
            '规格编码': 'spec_code',
            '成本价': 'cost_price',
            '市场': 'market',
            '档口': 'shop',
            '供应商': 'supplier',
            '平台': 'platform',
            '店铺': 'store',
            '商品ID': 'product_id',
            '商品': 'product_name',
            '商家编码': 'merchant_code',
            '简称': 'abbreviation',
            '图片URL': 'picture_url',
        }

        # 创建列索引到字段名的映射
        column_map = {}
        for idx, header in enumerate(headers):
            if header in verbose_name_map:
                column_map[idx] = verbose_name_map[header]

        print("列映射:", column_map)

        # 计算总行数
        total_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        task.total_rows = total_rows
        task.save()
        print(f"总行数: {total_rows}")

        # 重新打开工作簿进行数据处理
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        # 从第二行开始处理数据
        processed_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {}
            for idx, value in enumerate(row):
                if idx in column_map:
                    field_name = column_map[idx]
                    if field_name == 'cost_price':
                        data[field_name] = float(value) if value else 0
                    else:
                        data[field_name] = str(value) if value else ''

            # 打印前5行的数据用于调试
            # if processed_count < 5:
            #     print(f"第{processed_count+1}行数据: {data}")

            if data.get('spec_id'):
                SysShop.objects.update_or_create(
                    spec_id=data['spec_id'],
                    defaults={k: v for k, v in data.items() if k != 'spec_id'}
                )
                processed_count += 1

                # 每100行更新一次进度
                if processed_count % 100 == 0:
                    task.processed_rows = processed_count
                    task.save()
                    print(f"已处理: {processed_count}/{total_rows}")

        wb.close()

        # 更新任务状态为完成
        task.processed_rows = processed_count
        task.status = 'completed'
        task.save()

        print(f"任务完成: {task_id}, 共处理 {processed_count} 条数据")

    except Exception as e:
        print(f"任务失败: {task_id}, 错误: {str(e)}")
        task = ImportTask.objects.get(task_id=task_id)
        task.status = 'failed'
        task.error_message = str(e)
        task.save()


class GetShops(View):
    def get(self, request):
        try:
            # 获取分页参数
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 20))

            # 搜索参数
            keyword = request.GET.get('keyword', '')
            if keyword:
                shops = SysShop.objects.filter(
                    spec_id__icontains=keyword
                ) | SysShop.objects.filter(
                    spec_name__icontains=keyword
                ) | SysShop.objects.filter(
                    spec_alias__icontains=keyword
                ) | SysShop.objects.filter(
                    shop__icontains=keyword
                )
            else:
                shops = SysShop.objects.all()

            # 分页
            paginator = Paginator(shops, page_size)
            page_obj = paginator.get_page(page)

            # 转换为字典列表
            shops_list = []
            for shop in page_obj:
                shops_list.append({
                    'spec_id': shop.spec_id,
                    'spec_name': shop.spec_name,
                    'spec_alias': shop.spec_alias,
                    'spec_code': shop.spec_code,
                    'cost_price': shop.cost_price,
                    'market': shop.market,
                    'shop': shop.shop,
                    'supplier': shop.supplier,
                    'platform': shop.platform,
                    'store': shop.store,
                    'product_id': shop.product_id,
                    'product_name': shop.product_name,
                    'merchant_code': shop.merchant_code,
                    'abbreviation': shop.abbreviation,
                    'picture_url': shop.picture_url,
                })

            return JsonResponse({
                'code': 200,
                'data': {
                    'list': shops_list,
                    'total': paginator.count,
                    'page': page,
                    'page_size': page_size,
                    'total_pages': paginator.num_pages
                }
            })
        except Exception as e:
            return JsonResponse({'code': 500, 'errorInfo': str(e)})
