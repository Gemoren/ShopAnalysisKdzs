import os
import threading
import uuid
from datetime import datetime

import openpyxl
from django.conf import settings
from django.http import JsonResponse
from django.shortcuts import render
from .models import SysPromotion, PromotionSummaryMonthView
from user.models import ImportTask
from django.core.paginator import Paginator
from django.views import View


# Create your views here.
class UploadFile(View):
    def post(self, request):
        print("request:", request)
        file = request.FILES.get('file')
        print("file:", file)

        if file:
            file_name = file.name
            suffixName = file_name[file_name.rfind("."):]
            new_file_name = datetime.now().strftime('%Y%m%d%H%M%S') + suffixName
            file_path = str(settings.MEDIA_ROOT) + "\\promotionFile\\" + new_file_name

            if not os.path.exists(str(settings.MEDIA_ROOT) + "\\promotionFile\\"):
                os.makedirs(str(settings.MEDIA_ROOT) + "\\promotionFile\\")

            print("file_path:", file_path)

            try:
                # 保存文件
                with open(file_path, 'wb') as f:
                    for chunk in file.chunks():
                        f.write(chunk)

                # 创建导入任务
                task_id = str(uuid.uuid4())
                task = ImportTask.objects.create(
                    task_id=task_id,
                    file_name=file_name,
                    file_path=file_path,
                    status='pending'
                )

                # 启动后台线程处理文件
                thread = threading.Thread(target=process_excel_file, args=(task_id, file_path))
                thread.daemon = True
                thread.start()

                return JsonResponse({
                    'code': 200,
                    'message': '文件上传成功，正在后台处理',
                    'task_id': task_id,
                    'file_name': new_file_name
                })

            except Exception as e:
                return JsonResponse({'code': 500, 'errorInfo': f'文件上传失败: {str(e)}'})

        return JsonResponse({'code': 400, 'errorInfo': '请上传文件'})


# 异步处理Excel文件的函数
def process_excel_file(task_id, file_path):
    try:
        print(f"开始处理任务: {task_id}")

        # 更新任务状态为处理中
        task = ImportTask.objects.get(task_id=task_id)
        task.status = 'processing'
        task.save()

        # 打开表格
        wb = openpyxl.load_workbook(file_path)

        # 打印所有工作表名称
        print("所有工作表:", wb.sheetnames)

        # 查找以'商品_分天数据'开头的工作表
        target_sheet = None
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('商品_分天数据'):
                target_sheet = sheet_name
                print(f"找到目标工作表: {target_sheet}")
                break

        if not target_sheet:
            raise Exception('未找到以"商品_分天数据"开头的工作表')

        ws = wb[target_sheet]

        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        print("表头:", headers)

        # 先打印前3行数据用于调试
        print("\n前3行原始数据:")
        for i, row in enumerate(ws.iter_rows(max_row=3, values_only=True), 1):
            row_data = [str(cell) if cell is not None else '' for cell in row[:15]]
            print(f"第{i}行: {row_data}")

        # 表头在第1行
        header_row = 1

        # 获取表头
        headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        print(f"\n表头行 {header_row}: {len(headers)} 列")

        # 创建verbose_name到字段名的映射（根据实际Excel表头更新）
        verbose_name_map = {
            '日期': 'date',
            '商品ID': 'product_id',
            '商品名称': 'product_name',
            '推广场景': 'promotion_scene',
            '推广名称': 'promotion_name',
            '出价方式': 'bidding_method',
            '成交花费(元)': 'transaction_cost',
            '交易额(元)': 'transaction_amount',
            '实际投产比': 'actual_roi',
            '实际成交花费(元)': 'transaction_cost',  # 别名
            '总花费(元)': 'total_cost',
            '净交易额(元)': 'net_transaction_amount',
            '净实际投产比': 'net_actual_roi',
            '净成交笔数': 'net_transaction_count',
            '每笔净成交花费(元)': 'cost_per_net_transaction',
            '净交易额占比': 'net_transaction_ratio',
            '成交笔数': 'transaction_count',
            '每笔成交花费(元)': 'cost_per_transaction',
            '每笔成交金额(元)': 'amount_per_transaction',
            '直接交易额(元)': 'direct_transaction_amount',
            '间接交易额(元)': 'indirect_transaction_amount',
            '直接成交笔数': 'direct_transaction_count',
            '间接成交笔数': 'indirect_transaction_count',
            '每笔直接成交金额(元)': 'direct_amount_per_transaction',
            '每笔间接成交金额(元)': 'indirect_amount_per_transaction',
            '全站推广费比': 'site_promotion_ratio',
            '曝光量': 'exposure_count',
            '点击量': 'click_count',
            '询单花费(元)': 'inquiry_cost',
            '询单量': 'inquiry_count',
            '平均询单成本(元)': 'avg_inquiry_cost',
            '收藏花费(元)': 'favorite_cost',
            '收藏量': 'favorite_count',
            '平均收藏成本(元)': 'avg_favorite_cost',
            '关注花费(元)': 'follow_cost',
            '关注量': 'follow_count',
            '平均关注成本(元)': 'avg_follow_cost',
        }

        # 创建列索引到字段名的映射
        column_map = {}
        for idx, header in enumerate(headers):
            if header in verbose_name_map:
                column_map[idx] = verbose_name_map[header]

        print("列映射:", column_map)

        # 计算总行数（从表头下一行开始）
        data_start_row = header_row + 1
        total_rows = sum(1 for _ in ws.iter_rows(min_row=data_start_row, values_only=True))
        task.total_rows = total_rows
        task.save()
        print(f"总行数: {total_rows}")

        # 重新打开工作簿进行数据处理
        wb = openpyxl.load_workbook(file_path)

        # 查找以'商品_分天数据'开头的工作表
        target_sheet = None
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('商品_分天数据'):
                target_sheet = sheet_name
                print(f"找到目标工作表: {target_sheet}")
                break

        if not target_sheet:
            raise Exception('未找到以"商品_分天数据"开头的工作表')

        ws = wb[target_sheet]

        # 定义验证日期格式的函数
        def is_valid_date(value):
            if value is None:
                return False
            # 如果是 datetime 类型，直接返回 True
            if isinstance(value, datetime):
                return True
            # 如果是字符串，尝试解析
            if isinstance(value, str):
                try:
                    # 尝试常见的日期格式
                    datetime.strptime(value, '%Y-%m-%d')
                    return True
                except ValueError:
                    try:
                        datetime.strptime(value, '%Y/%m/%d')
                        return True
                    except ValueError:
                        try:
                            datetime.strptime(value, '%Y年%m月%d日')
                            return True
                        except ValueError:
                            return False
            # 如果是数字类型，尝试转换为日期
            try:
                from datetime import timedelta
                date = datetime(1899, 12, 30) + timedelta(days=int(value))
                return True
            except:
                return False
            return False

        # 从表头下一行开始处理数据
        processed_count = 0
        skipped_count = 0
        for i, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), data_start_row):
            # 获取日期列的值（第一列）
            row_date = row[0] if row and len(row) > 0 else None

            # 跳过日期为 "总计" 的行
            if row_date == '总计' or (isinstance(row_date, str) and '总计' in row_date):
                print(f"跳过总计行: 第 {i} 行")
                skipped_count += 1
                continue

            # 验证日期格式，如果不是有效日期则跳过
            if not is_valid_date(row_date):
                print(f"跳过无效日期行: 第 {i} 行, 日期值: {row_date}")
                skipped_count += 1
                continue

            data = {}
            for idx, value in enumerate(row):
                if idx in column_map:
                    field_name = column_map[idx]
                    if field_name == 'date':
                        # 跳过无效的日期值
                        if not value or isinstance(value, str) and value in ['', ' ', '总计']:
                            continue
                        data[field_name] = value
                    elif field_name in ['product_id', 'product_name', 'promotion_scene', 'promotion_name',
                                        'bidding_method']:
                        data[field_name] = str(value) if value else ''
                    elif field_name in ['total_cost', 'transaction_cost', 'transaction_amount',
                                        'actual_roi', 'net_actual_roi', 'net_transaction_amount',
                                        'cost_per_net_transaction', 'net_transaction_ratio',
                                        'cost_per_transaction', 'amount_per_transaction',
                                        'direct_transaction_amount', 'indirect_transaction_amount',
                                        'direct_amount_per_transaction', 'indirect_amount_per_transaction',
                                        'site_promotion_ratio', 'inquiry_cost', 'avg_inquiry_cost',
                                        'favorite_cost', 'avg_favorite_cost', 'follow_cost', 'avg_follow_cost']:
                        # 处理百分号格式，如 "100.00%" -> 100.00
                        if value and isinstance(value, str) and '%' in value:
                            data[field_name] = float(value.replace('%', ''))
                        elif isinstance(value, str) and value.strip() == '-':
                            # 处理 "-" 这样的特殊字符，转换为 0
                            data[field_name] = 0
                        else:
                            data[field_name] = float(value) if value else 0
                    elif field_name in ['net_transaction_count', 'transaction_count', 'direct_transaction_count',
                                        'indirect_transaction_count', 'exposure_count', 'click_count',
                                        'inquiry_count', 'favorite_count', 'follow_count']:
                        if isinstance(value, str) and value.strip() == '-':
                            # 处理 "-" 这样的特殊字符，转换为 0
                            data[field_name] = 0
                        else:
                            data[field_name] = int(value) if value else 0

            # 检查是否有有效数据
            if data.get('date') and data.get('product_id'):
                SysPromotion.objects.update_or_create(
                    date=data['date'],
                    product_id=data['product_id'],
                    defaults={k: v for k, v in data.items() if k not in ['date', 'product_id']}
                )
                processed_count += 1

                # 每100行更新一次进度
                if processed_count % 100 == 0:
                    task.processed_rows = processed_count
                    task.save()
                    print(f"已处理: {processed_count}/{total_rows} (当前行: {i})")

        wb.close()

        # 更新任务状态为完成
        task.processed_rows = processed_count
        task.status = 'completed'
        task.save()

        print(f"任务完成: {task_id}, 共处理 {processed_count} 条数据, 跳过 {skipped_count} 条无效数据")

    except Exception as e:
        print(f"任务失败: {task_id}, 错误: {str(e)}")
        task = ImportTask.objects.get(task_id=task_id)
        task.status = 'failed'
        task.error_message = str(e)
        task.save()


class GetPromotions(View):
    def get(self, request):
        try:
            # 获取分页参数
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 20))

            # 搜索参数
            keyword = request.GET.get('keyword', '')
            if keyword:
                promotions = SysPromotion.objects.filter(
                    product_id__icontains=keyword
                ) | SysPromotion.objects.filter(
                    product_name__icontains=keyword
                ) | SysPromotion.objects.filter(
                    promotion_name__icontains=keyword
                )
            else:
                promotions = SysPromotion.objects.all()

            # 分页
            paginator = Paginator(promotions, page_size)
            page_obj = paginator.get_page(page)

            # 转换为字典列表
            promotions_list = []
            for promotion in page_obj:
                promotions_list.append({
                    'date': promotion.date.strftime('%Y-%m-%d') if promotion.date else '',
                    'product_id': promotion.product_id,
                    'product_name': promotion.product_name,
                    'promotion_scene': promotion.promotion_scene,
                    'promotion_name': promotion.promotion_name,
                    'bidding_method': promotion.bidding_method,
                    'total_cost': promotion.total_cost,
                    'transaction_cost': promotion.transaction_cost,
                    'transaction_amount': promotion.transaction_amount,
                    'actual_roi': promotion.actual_roi,
                    'net_actual_roi': promotion.net_actual_roi,
                    'net_transaction_amount': promotion.net_transaction_amount,
                    'net_transaction_count': promotion.net_transaction_count,
                    'cost_per_net_transaction': promotion.cost_per_net_transaction,
                    'net_transaction_ratio': promotion.net_transaction_ratio,
                    'transaction_count': promotion.transaction_count,
                    'cost_per_transaction': promotion.cost_per_transaction,
                    'amount_per_transaction': promotion.amount_per_transaction,
                    'direct_transaction_amount': promotion.direct_transaction_amount,
                    'indirect_transaction_amount': promotion.indirect_transaction_amount,
                    'direct_transaction_count': promotion.direct_transaction_count,
                    'indirect_transaction_count': promotion.indirect_transaction_count,
                    'direct_amount_per_transaction': promotion.direct_amount_per_transaction,
                    'indirect_amount_per_transaction': promotion.indirect_amount_per_transaction,
                    'site_promotion_ratio': promotion.site_promotion_ratio,
                    'exposure_count': promotion.exposure_count,
                    'click_count': promotion.click_count,
                    'inquiry_cost': promotion.inquiry_cost,
                    'inquiry_count': promotion.inquiry_count,
                    'avg_inquiry_cost': promotion.avg_inquiry_cost,
                    'favorite_cost': promotion.favorite_cost,
                    'favorite_count': promotion.favorite_count,
                    'avg_favorite_cost': promotion.avg_favorite_cost,
                    'follow_cost': promotion.follow_cost,
                    'follow_count': promotion.follow_count,
                    'avg_follow_cost': promotion.avg_follow_cost,
                })

            return JsonResponse({
                'code': 200,
                'data': {
                    'list': promotions_list,
                    'total': paginator.count,
                    'page': page,
                    'page_size': page_size,
                    'total_pages': paginator.num_pages
                }
            })
        except Exception as e:
            return JsonResponse({'code': 500, 'errorInfo': str(e)})


# 按年月和店铺聚合总花费(元)
class GetPromotionsByMonth(View):
    def get(self, request):
        try:
            # 获取时间范围参数
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            # 从 PromotionSummaryMonthView 视图中获取数据，使用 values 避免访问不存在的 id 字段
            promotions = PromotionSummaryMonthView.objects.all()

            print(promotions)

            # 按时间范围过滤
            if start_date:
                promotions = promotions.filter(month__gte=start_date)
            if end_date:
                promotions = promotions.filter(month__lte=end_date)

            print(promotions)
            # 使用 values() 获取字典形式的数据，避免 ORM 对象访问
            promotions_data = promotions.values('store', 'month', 'total_cost')

            # 转换为字典列表
            promotions_list = []
            for promotion in promotions_data:
                promotions_list.append({
                    'month_year': promotion['month'] + '-01',  # 转换为日期格式
                    'store_name': promotion['store'] if promotion['store'] else '未分类',
                    'total_cost_sum': float(promotion['total_cost']) if promotion['total_cost'] else 0,
                })

            return JsonResponse({
                'code': 200,
                'data': {
                    'list': promotions_list
                }
            })
        except Exception as e:
            print(e)
            return JsonResponse({'code': 500, 'errorInfo': str(e)})
