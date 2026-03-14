import os
import threading
import uuid
from datetime import datetime
from django.utils import timezone

import openpyxl
from django.conf import settings
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.http import JsonResponse
from django.shortcuts import render
from .models import SysOrder
from user.models import ImportTask
from django.core.paginator import Paginator
from django.views import View


# Create your views here.
class UploadFile(View):
    def post(self, request):
        print("request:", request)
        file = request.FILES.get('file')
        print("file:", file)

        if file:
            file_name = file.name
            suffixName = file_name[file_name.rfind("."):]
            new_file_name = datetime.now().strftime('%Y%m%d%H%M%S') + suffixName
            file_path = str(settings.MEDIA_ROOT) + "\\orderFile\\" + new_file_name

            if not os.path.exists(str(settings.MEDIA_ROOT) + "\\orderFile\\"):
                os.makedirs(str(settings.MEDIA_ROOT) + "\\orderFile\\")

            print("file_path:", file_path)

            try:
                # 保存文件
                with open(file_path, 'wb') as f:
                    for chunk in file.chunks():
                        f.write(chunk)

                # 创建导入任务
                task_id = str(uuid.uuid4())
                task = ImportTask.objects.create(
                    task_id=task_id,
                    file_name=file_name,
                    file_path=file_path,
                    status='pending'
                )

                # 启动后台线程处理文件
                thread = threading.Thread(target=process_excel_file, args=(task_id, file_path))
                thread.daemon = True
                thread.start()

                return JsonResponse({
                    'code': 200,
                    'message': '文件上传成功，正在后台处理',
                    'task_id': task_id,
                    'file_name': new_file_name
                })

            except Exception as e:
                return JsonResponse({'code': 500, 'errorInfo': f'文件上传失败: {str(e)}'})

        return JsonResponse({'code': 400, 'errorInfo': '请上传文件'})


# 异步处理Excel文件的函数
def process_excel_file(task_id, file_path):
    try:
        print(f"开始处理任务: {task_id}")

        # 更新任务状态为处理中
        task = ImportTask.objects.get(task_id=task_id)
        task.status = 'processing'
        task.save()

        # 打开表格
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        # 获取表头（第一行）
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        print("表头:", headers)

        # 创建verbose_name到字段名的映射
        verbose_name_map = {
            '订单编号': 'order_no',
            '平台': 'platform',
            '店铺名称': 'store_name',
            '下单时间': 'order_time',
            '订单状态': 'order_status',
            '快递公司': 'courier_company',
            '运单号': 'tracking_no',
            '商品总价': 'product_total_price',
            '宝贝数量': 'product_count',
            '宝贝种类': 'product_types',
            '实付金额': 'actual_payment',
            '实收金额': 'actual_received',
            '店铺优惠金额': 'store_discount',
            '平台优惠金额': 'platform_discount',
            '系统单号': 'system_no',
            '平台商品ID': 'platform_product_id',
            '平台skuID': 'platform_sku_id',
            '规格名称': 'spec_name',
            '别名': 'alias',
        }

        # 创建列索引到字段名的映射
        column_map = {}
        for idx, header in enumerate(headers):
            if header in verbose_name_map:
                column_map[idx] = verbose_name_map[header]

        print("列映射:", column_map)

        # 计算总行数
        total_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        task.total_rows = total_rows
        task.save()
        print(f"总行数: {total_rows}")

        # 重新打开工作簿进行数据处理
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        # 从第二行开始处理数据
        processed_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {}
            for idx, value in enumerate(row):
                if idx in column_map:
                    field_name = column_map[idx]
                    if field_name == 'order_time' and value:
                        data[field_name] = value
                    elif field_name in ['product_total_price', 'actual_payment', 'actual_received',
                                        'store_discount', 'platform_discount']:
                        data[field_name] = float(value) if value else 0
                    elif field_name in ['product_count', 'product_types']:
                        data[field_name] = int(value) if value else 0
                    else:
                        data[field_name] = str(value) if value else ''

            if data.get('order_no'):
                SysOrder.objects.update_or_create(
                    order_no=data['order_no'],
                    defaults={k: v for k, v in data.items() if k != 'order_no'}
                )
                processed_count += 1

                # 每100行更新一次进度
                if processed_count % 100 == 0:
                    task.processed_rows = processed_count
                    task.save()
                    print(f"已处理: {processed_count}/{total_rows}")

        wb.close()

        # 更新任务状态为完成
        task.processed_rows = processed_count
        task.status = 'completed'
        task.save()

        print(f"任务完成: {task_id}, 共处理 {processed_count} 条数据")

    except Exception as e:
        print(f"任务失败: {task_id}, 错误: {str(e)}")
        task = ImportTask.objects.get(task_id=task_id)
        task.status = 'failed'
        task.error_message = str(e)
        task.save()


class GetOrders(View):
    def get(self, request):
        try:
            # 获取分页参数
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 20))

            # 搜索参数
            keyword = request.GET.get('keyword', '')
            if keyword:
                orders = SysOrder.objects.filter(
                    order_no__icontains=keyword
                ) | SysOrder.objects.filter(
                    platform__icontains=keyword
                ) | SysOrder.objects.filter(
                    store_name__icontains=keyword
                ) | SysOrder.objects.filter(
                    spec_name__icontains=keyword
                )
            else:
                orders = SysOrder.objects.all()

            # 分页
            paginator = Paginator(orders, page_size)
            page_obj = paginator.get_page(page)

            # 转换为字典列表
            orders_list = []
            for order in page_obj:
                orders_list.append({
                    'order_no': order.order_no,
                    'platform': order.platform,
                    'store_name': order.store_name,
                    'order_time': order.order_time.strftime('%Y-%m-%d %H:%M:%S') if order.order_time else '',
                    'order_status': order.order_status,
                    'courier_company': order.courier_company,
                    'tracking_no': order.tracking_no,
                    'product_total_price': order.product_total_price,
                    'product_count': order.product_count,
                    'product_types': order.product_types,
                    'actual_payment': order.actual_payment,
                    'actual_received': order.actual_received,
                    'store_discount': order.store_discount,
                    'platform_discount': order.platform_discount,
                    'system_no': order.system_no,
                    'platform_product_id': order.platform_product_id,
                    'platform_sku_id': order.platform_sku_id,
                    'spec_name': order.spec_name,
                    'alias': order.alias,
                })

            return JsonResponse({
                'code': 200,
                'data': {
                    'list': orders_list,
                    'total': paginator.count,
                    'page': page,
                    'page_size': page_size,
                    'total_pages': paginator.num_pages
                }
            })
        except Exception as e:
            return JsonResponse({'code': 500, 'errorInfo': str(e)})


# '下单时间'字段按年月、'店铺名称'聚合'实收金额'，仅要'订单状态'为'交易成功'的数据
class GetOrdersByMonth(View):
    def get(self, request):
        try:
            # 获取时间范围参数
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            # 筛选订单状态为'交易成功'的数据
            orders = SysOrder.objects.filter(order_status='交易成功')

            # 根据时间范围过滤
            if start_date:
                try:
                    start_datetime = datetime.strptime(start_date, '%Y-%m')
                    # 转换为当前时区
                    if timezone.is_naive(start_datetime):
                        start_datetime = timezone.make_aware(start_datetime)
                    orders = orders.filter(order_time__gte=start_datetime)
                except ValueError:
                    pass

            if end_date:
                try:
                    # 结束日期要包含该月的所有时间，所以设置为下月的第一天减1秒
                    end_datetime = datetime.strptime(end_date, '%Y-%m')
                    # 转换为当前时区
                    if timezone.is_naive(end_datetime):
                        end_datetime = timezone.make_aware(end_datetime)
                    # 计算下个月的第一天
                    if end_datetime.month == 12:
                        end_datetime = end_datetime.replace(year=end_datetime.year + 1, month=1)
                    else:
                        end_datetime = end_datetime.replace(month=end_datetime.month + 1)
                    orders = orders.filter(order_time__lt=end_datetime)
                except ValueError:
                    pass

            # 按年月和店铺名称分组，聚合实收金额
            orders_by_month = orders.annotate(
                month_year=TruncMonth('order_time')
            ).values('month_year', 'store_name').annotate(
                total_received=Sum('actual_received'))
            # 按年月和店铺名称排序
            orders_by_month = orders_by_month.order_by('month_year', 'store_name')
            return JsonResponse({
                'code': 200,
                'data': {
                    'list': list(orders_by_month),
                }
            })
        except Exception as e:
            return JsonResponse({'code': 500, 'errorInfo': str(e)})
